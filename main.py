import tkinter as tk
from tkinter import ttk, filedialog, messagebox, simpledialog
import pandas as pd
import os
import sys

from database import Database
from decoder import decode_codes

# --- Color theme ---
BG = '#f0f4f8'
CARD_BG = '#ffffff'
PRIMARY = '#2563eb'
PRIMARY_HOVER = '#1d4ed8'
SUCCESS = '#16a34a'
WARNING = '#e67e22'
TEXT = '#1e293b'
TEXT_LIGHT = '#64748b'
DANGER = '#dc2626'

APP_VERSION = 'v1.2 — 2026'
APP_TITLE = 'ORVA SSKU Digit Code Decoder'
FOOTER_TEXT = f'ORAVA (Pvt) Ltd  |  Developed by Shehan Nirmana  |  {APP_VERSION}'


def resource_path(filename):
    """Return absolute path to a bundled resource (works for dev and PyInstaller)."""
    if getattr(sys, 'frozen', False):
        base = getattr(sys, '_MEIPASS', os.path.dirname(sys.executable))
    else:
        base = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base, filename)


def load_image(filename, size=None):
    """Load an image using PIL if available, else fall back to tk.PhotoImage."""
    path = resource_path(filename)
    if not os.path.exists(path):
        return None
    try:
        from PIL import Image, ImageTk
        img = Image.open(path)
        if size:
            img = img.resize(size, Image.LANCZOS)
        return ImageTk.PhotoImage(img)
    except Exception:
        try:
            return tk.PhotoImage(file=path)
        except Exception:
            return None


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(APP_TITLE)
        self.geometry('880x820')
        self.resizable(False, False)
        self.configure(bg=BG)

        # Window icon
        self._icon_image = load_image('icon.png')
        if self._icon_image is not None:
            try:
                self.iconphoto(True, self._icon_image)
            except Exception:
                pass

        # Shared logo image — preserve 1200x424 (~3:1) aspect ratio
        self.logo_image = load_image('logo.png', size=(180, 64))

        self.db = Database()
        self.loaded_codes = []
        self.decoded_results = []
        self.skipped_codes = []

        # Container for screens
        container = tk.Frame(self, bg=BG)
        container.pack(fill='both', expand=True)
        container.grid_rowconfigure(0, weight=1)
        container.grid_columnconfigure(0, weight=1)

        self.frames = {}
        for Screen in (LibraryScreen, CodeUploadScreen, ExportScreen):
            frame = Screen(container, self)
            self.frames[Screen.__name__] = frame
            frame.grid(row=0, column=0, sticky='nsew')

        self.show('LibraryScreen')
        self.protocol('WM_DELETE_WINDOW', self.on_close)

    def show(self, name):
        frame = self.frames[name]
        frame.tkraise()
        if hasattr(frame, 'on_show'):
            frame.on_show()

    def on_close(self):
        self.db.close()
        self.destroy()


# --- Reusable widgets ---

def make_header(parent, app):
    """Top bar with company logo on every screen."""
    header = tk.Frame(parent, bg=BG)
    header.pack(side='top', fill='x', pady=(10, 0))
    if app.logo_image is not None:
        tk.Label(header, image=app.logo_image, bg=BG).pack()
    else:
        tk.Label(header, text='ORAVA', font=('Segoe UI', 16, 'bold'),
                 bg=BG, fg=PRIMARY).pack()
    return header


def make_footer(parent):
    """Bottom footer shown on every screen."""
    footer = tk.Label(parent, text=FOOTER_TEXT, font=('Segoe UI', 8),
                      bg=BG, fg=TEXT_LIGHT)
    footer.pack(side='bottom', pady=6)
    return footer


def make_title(parent, text):
    lbl = tk.Label(parent, text=text, font=('Segoe UI', 18, 'bold'),
                   bg=BG, fg=TEXT)
    lbl.pack(pady=(10, 4))


def make_subtitle(parent, text):
    lbl = tk.Label(parent, text=text, font=('Segoe UI', 10),
                   bg=BG, fg=TEXT_LIGHT)
    lbl.pack(pady=(0, 12))


def make_button(parent, text, command, color=PRIMARY, width=18):
    btn = tk.Button(parent, text=text, command=command,
                    font=('Segoe UI', 10, 'bold'), fg='white', bg=color,
                    activebackground=PRIMARY_HOVER, activeforeground='white',
                    relief='flat', cursor='hand2', width=width, pady=6)
    return btn


# --- Screen 1: Library Management ---

LIBRARY_CONFIG = [
    ('stone_names', '1-4 Digit', 'Stone Name Library',
     '1-4 degit', 'Stone name'),
    ('polishing_types', '5th Digit', 'Polishing Type Library',
     '5 degit', 'Polishing Type'),
    ('shapes', '6-7 Digit', 'Shape Library',
     '6-7 degit', 'Shape'),
    ('colours', '8-9 Digit', 'Colour Code Library',
     '8-9 degit', 'Colour'),
    ('origins', '10-11 Digit', 'Origin Library',
     '10-11 degit', 'Origin'),
]


class LibraryScreen(tk.Frame):
    def __init__(self, parent, app):
        super().__init__(parent, bg=BG)
        self.app = app
        self.status_labels = {}

        make_header(self, app)
        make_footer(self)

        make_title(self, 'Library Management')
        make_subtitle(self, 'Upload master data files. Libraries are saved permanently.')

        cards_frame = tk.Frame(self, bg=BG)
        cards_frame.pack(fill='both', expand=True, padx=30, pady=(0, 5))

        for i, (table, digit, label, col_code, col_name) in enumerate(LIBRARY_CONFIG):
            card = tk.Frame(cards_frame, bg=CARD_BG, relief='groove', bd=1)
            card.pack(fill='x', pady=4, ipady=4)

            left = tk.Frame(card, bg=CARD_BG)
            left.pack(side='left', fill='x', expand=True, padx=12)

            tk.Label(left, text=f'{label}  ({digit})',
                     font=('Segoe UI', 10, 'bold'), bg=CARD_BG, fg=TEXT
                     ).pack(anchor='w')

            status = tk.Label(left, text='', font=('Segoe UI', 9),
                              bg=CARD_BG, fg=TEXT_LIGHT)
            status.pack(anchor='w')
            self.status_labels[table] = status

            btns = tk.Frame(card, bg=CARD_BG)
            btns.pack(side='right', padx=8, pady=2)
            make_button(btns, 'Upload',
                        lambda t=table, cc=col_code, cn=col_name:
                        self.upload_library(t, cc, cn),
                        width=8).pack(side='left', padx=2)
            make_button(btns, 'Edit',
                        lambda t=table, lbl=label:
                        self.open_edit_popup(t, lbl),
                        color=TEXT_LIGHT, width=7).pack(side='left', padx=2)
            make_button(btns, 'Download',
                        lambda t=table, lbl=label:
                        self.download_library(t, lbl),
                        color=SUCCESS, width=9).pack(side='left', padx=2)
            make_button(btns, 'Delete All',
                        lambda t=table, lbl=label:
                        self.delete_all_library(t, lbl),
                        color=DANGER, width=10).pack(side='left', padx=2)

        btn_frame = tk.Frame(self, bg=BG)
        btn_frame.pack(side='bottom', pady=12)
        make_button(btn_frame, 'Next  ➜', lambda: self.app.show('CodeUploadScreen'),
                    width=14).pack(side='left', padx=6)

    def on_show(self):
        for table, _, _, _, _ in LIBRARY_CONFIG:
            count = self.app.db.get_count(table)
            self.status_labels[table].config(
                text=f'{count} entries loaded',
                fg=SUCCESS if count > 0 else TEXT_LIGHT)

    def upload_library(self, table, col_code, col_name):
        path = filedialog.askopenfilename(
            title=f'Select {table} library file',
            filetypes=[('Excel files', '*.xlsx *.xls'), ('All files', '*.*')])
        if not path:
            return
        try:
            df = pd.read_excel(path)
            code_col = self._find_column(df, col_code)
            name_col = self._find_column(df, col_name)
            if code_col is None or name_col is None:
                if len(df.columns) >= 2:
                    code_col = df.columns[0]
                    name_col = df.columns[1]
                else:
                    messagebox.showerror('Error',
                                         f'Could not find columns. Expected "{col_code}" and "{col_name}".')
                    return

            entries = list(zip(df[code_col].astype(str), df[name_col].astype(str)))
            added, skipped = self.app.db.add_entries(table, entries)
            total = self.app.db.get_count(table)
            self.status_labels[table].config(
                text=f'{total} entries loaded  (+{added} new, {skipped} existing)',
                fg=SUCCESS)
            messagebox.showinfo('Success',
                                f'Added {added} new entries.\n{skipped} duplicates skipped.\nTotal: {total}')
        except Exception as e:
            messagebox.showerror('Error', f'Failed to read file:\n{e}')

    def _find_column(self, df, target):
        target_lower = target.lower()
        for col in df.columns:
            if str(col).lower().strip() == target_lower:
                return col
        for col in df.columns:
            if target_lower in str(col).lower().strip():
                return col
        return None

    def download_library(self, table, label):
        if self.app.db.get_count(table) == 0:
            messagebox.showinfo('Empty', f'{label} has no entries to export.')
            return
        path = filedialog.asksaveasfilename(
            title=f'Save {label}',
            defaultextension='.xlsx',
            filetypes=[('Excel files', '*.xlsx')],
            initialfile=f'{table}.xlsx')
        if not path:
            return
        try:
            entries = self.app.db.get_all_entries(table)
            df = pd.DataFrame(entries, columns=['Code', 'Meaning'])
            df.to_excel(path, index=False)
            messagebox.showinfo('Success',
                                f'Exported {len(entries)} entries to:\n{path}')
        except Exception as e:
            messagebox.showerror('Error', f'Export failed:\n{e}')

    def delete_all_library(self, table, label):
        count = self.app.db.get_count(table)
        if count == 0:
            messagebox.showinfo('Empty', f'{label} is already empty.')
            return
        if messagebox.askyesno(
                'Confirm Delete All',
                f'This will permanently delete all {count} entries from "{label}".\n\nThis cannot be undone. Continue?',
                icon='warning'):
            self.app.db.delete_all(table)
            self.on_show()
            messagebox.showinfo('Done', f'{label} cleared.')

    def open_edit_popup(self, table, label):
        win = tk.Toplevel(self)
        win.title(f'Edit — {label}')
        win.geometry('640x560')
        win.configure(bg=BG)
        win.transient(self.winfo_toplevel())
        win.grab_set()

        tk.Label(win, text=f'Edit {label}',
                 font=('Segoe UI', 14, 'bold'), bg=BG, fg=TEXT
                 ).pack(pady=(10, 4))

        # --- Add new entry form ---
        form = tk.Frame(win, bg=CARD_BG, relief='groove', bd=1)
        form.pack(fill='x', padx=12, pady=6, ipady=8, ipadx=8)
        tk.Label(form, text='Add New Entry',
                 font=('Segoe UI', 10, 'bold'), bg=CARD_BG, fg=TEXT
                 ).grid(row=0, column=0, columnspan=5, sticky='w',
                        padx=6, pady=(4, 6))
        tk.Label(form, text='Code:', font=('Segoe UI', 10),
                 bg=CARD_BG, fg=TEXT).grid(row=1, column=0, padx=6, pady=4, sticky='e')
        code_entry = tk.Entry(form, font=('Segoe UI', 10), width=12)
        code_entry.grid(row=1, column=1, padx=4, pady=4)
        tk.Label(form, text='Meaning:', font=('Segoe UI', 10),
                 bg=CARD_BG, fg=TEXT).grid(row=1, column=2, padx=6, pady=4, sticky='e')
        name_entry = tk.Entry(form, font=('Segoe UI', 10), width=28)
        name_entry.grid(row=1, column=3, padx=4, pady=4)

        # --- Search bar ---
        search_state = {'term': ''}

        search_frame = tk.Frame(win, bg=CARD_BG, relief='groove', bd=1)
        search_frame.pack(fill='x', padx=12, pady=(0, 4), ipady=6, ipadx=8)
        tk.Label(search_frame, text='Search:', font=('Segoe UI', 10, 'bold'),
                 bg=CARD_BG, fg=TEXT).pack(side='left', padx=(6, 4), pady=4)
        search_entry = tk.Entry(search_frame, font=('Segoe UI', 10), width=30)
        search_entry.pack(side='left', padx=4, pady=4)

        def do_search():
            search_state['term'] = search_entry.get().strip().lower()
            refresh()

        def do_clear():
            search_entry.delete(0, 'end')
            search_state['term'] = ''
            refresh()

        make_button(search_frame, 'Search', do_search,
                    color=PRIMARY, width=8).pack(side='left', padx=4, pady=4)
        make_button(search_frame, 'Clear', do_clear,
                    color=TEXT_LIGHT, width=8).pack(side='left', padx=4, pady=4)
        search_entry.bind('<Return>', lambda e: do_search())

        # --- Scrollable table ---
        table_frame = tk.Frame(win, bg=BG)
        table_frame.pack(fill='both', expand=True, padx=12, pady=6)

        canvas = tk.Canvas(table_frame, bg=CARD_BG, highlightthickness=0)
        vsb = ttk.Scrollbar(table_frame, orient='vertical', command=canvas.yview)
        inner = tk.Frame(canvas, bg=CARD_BG)
        inner.bind('<Configure>',
                   lambda e: canvas.configure(scrollregion=canvas.bbox('all')))
        canvas.create_window((0, 0), window=inner, anchor='nw')
        canvas.configure(yscrollcommand=vsb.set)
        canvas.pack(side='left', fill='both', expand=True)
        vsb.pack(side='right', fill='y')

        def _on_mousewheel(e):
            canvas.yview_scroll(int(-1 * (e.delta / 120)), 'units')
        canvas.bind('<Enter>',
                    lambda e: canvas.bind_all('<MouseWheel>', _on_mousewheel))
        canvas.bind('<Leave>',
                    lambda e: canvas.unbind_all('<MouseWheel>'))

        def refresh():
            for w in inner.winfo_children():
                w.destroy()
            hdr = tk.Frame(inner, bg=PRIMARY)
            hdr.pack(fill='x')
            tk.Label(hdr, text='Code', font=('Segoe UI', 10, 'bold'),
                     bg=PRIMARY, fg='white', width=12, anchor='w'
                     ).pack(side='left', padx=8, pady=5)
            tk.Label(hdr, text='Meaning', font=('Segoe UI', 10, 'bold'),
                     bg=PRIMARY, fg='white', anchor='w'
                     ).pack(side='left', padx=8, pady=5, fill='x', expand=True)
            tk.Label(hdr, text='Actions', font=('Segoe UI', 10, 'bold'),
                     bg=PRIMARY, fg='white', width=16, anchor='center'
                     ).pack(side='right', padx=8, pady=5)

            entries = self.app.db.get_all_entries(table)
            term = search_state['term']
            if term:
                entries = [(c, n) for c, n in entries
                           if term in str(c).lower() or term in str(n).lower()]
            if not entries:
                msg = ('(no matches — try a different search or clear)'
                       if term else '(no entries — add one above)')
                tk.Label(inner, text=msg,
                         font=('Segoe UI', 10, 'italic'),
                         bg=CARD_BG, fg=TEXT_LIGHT
                         ).pack(pady=20)
            for code, name in entries:
                row = tk.Frame(inner, bg=CARD_BG)
                row.pack(fill='x')
                tk.Label(row, text=code, font=('Consolas', 10),
                         bg=CARD_BG, fg=TEXT, width=12, anchor='w'
                         ).pack(side='left', padx=8, pady=3)
                tk.Label(row, text=name, font=('Segoe UI', 10),
                         bg=CARD_BG, fg=TEXT, anchor='w'
                         ).pack(side='left', padx=8, pady=3,
                                fill='x', expand=True)
                tk.Button(row, text='Delete',
                          command=lambda c=code: delete_one(c),
                          font=('Segoe UI', 9, 'bold'), fg='white',
                          bg=DANGER, activebackground=DANGER,
                          relief='flat', cursor='hand2', width=7
                          ).pack(side='right', padx=4, pady=2)
                tk.Button(row, text='Edit',
                          command=lambda c=code, n=name: edit_one(c, n),
                          font=('Segoe UI', 9, 'bold'), fg='white',
                          bg=PRIMARY, activebackground=PRIMARY_HOVER,
                          relief='flat', cursor='hand2', width=7
                          ).pack(side='right', padx=4, pady=2)
                tk.Frame(inner, bg='#e2e8f0', height=1).pack(fill='x')

        def add_new():
            code = code_entry.get().strip()
            name = name_entry.get().strip()
            if not code or not name:
                messagebox.showwarning(
                    'Missing field',
                    'Both Code and Meaning are required.', parent=win)
                return
            added, _ = self.app.db.add_entries(table, [(code, name)])
            if added:
                code_entry.delete(0, 'end')
                name_entry.delete(0, 'end')
                code_entry.focus_set()
                refresh()
            else:
                messagebox.showwarning(
                    'Duplicate',
                    f'Code "{code}" already exists.', parent=win)

        make_button(form, 'Add', add_new, color=SUCCESS, width=8
                    ).grid(row=1, column=4, padx=8, pady=4)

        def delete_one(code):
            if messagebox.askyesno(
                    'Confirm Delete',
                    f'Delete entry "{code}"?', parent=win):
                self.app.db.delete_entry(table, code)
                refresh()

        def edit_one(code, current_name):
            new_name = simpledialog.askstring(
                'Edit Meaning',
                f'New meaning for code "{code}":',
                initialvalue=current_name, parent=win)
            if new_name is not None and new_name.strip():
                self.app.db.update_entry(table, code, new_name.strip())
                refresh()

        def on_close():
            try:
                canvas.unbind_all('<MouseWheel>')
            except Exception:
                pass
            win.grab_release()
            win.destroy()
            self.on_show()

        win.protocol('WM_DELETE_WINDOW', on_close)

        btn_bar = tk.Frame(win, bg=BG)
        btn_bar.pack(fill='x', pady=8)
        make_button(btn_bar, 'Close', on_close,
                    color=TEXT_LIGHT, width=12).pack()

        refresh()


# --- Screen 2: Code Upload ---

class CodeUploadScreen(tk.Frame):
    def __init__(self, parent, app):
        super().__init__(parent, bg=BG)
        self.app = app

        make_header(self, app)
        make_footer(self)

        make_title(self, 'Upload Codes')
        make_subtitle(self, 'Upload an Excel file containing 11-digit stone codes.')

        upload_frame = tk.Frame(self, bg=CARD_BG, relief='groove', bd=1)
        upload_frame.pack(padx=40, pady=10, fill='x', ipady=20)

        self.file_label = tk.Label(upload_frame, text='No file selected',
                                   font=('Segoe UI', 10), bg=CARD_BG, fg=TEXT_LIGHT)
        self.file_label.pack(pady=(15, 8))

        make_button(upload_frame, 'Browse Excel File',
                    self.upload_codes, width=20).pack(pady=(0, 5))

        self.code_count_label = tk.Label(upload_frame, text='',
                                         font=('Segoe UI', 10, 'bold'),
                                         bg=CARD_BG, fg=SUCCESS)
        self.code_count_label.pack(pady=(5, 10))

        preview_frame = tk.Frame(self, bg=BG)
        preview_frame.pack(fill='both', expand=True, padx=40, pady=5)

        tk.Label(preview_frame, text='Code Preview:',
                 font=('Segoe UI', 10, 'bold'), bg=BG, fg=TEXT).pack(anchor='w')

        self.preview_text = tk.Text(preview_frame, height=12, width=50,
                                    font=('Consolas', 10), state='disabled',
                                    bg=CARD_BG, relief='groove', bd=1)
        self.preview_text.pack(fill='both', expand=True, pady=4)

        btn_frame = tk.Frame(self, bg=BG)
        btn_frame.pack(pady=12)
        make_button(btn_frame, '⬅  Libraries',
                    lambda: self.app.show('LibraryScreen'),
                    color=TEXT_LIGHT, width=14).pack(side='left', padx=6)
        make_button(btn_frame, 'Process & Next  ➜',
                    self.process_and_next, width=16).pack(side='left', padx=6)

    def upload_codes(self):
        path = filedialog.askopenfilename(
            title='Select codes Excel file',
            filetypes=[('Excel files', '*.xlsx *.xls'), ('All files', '*.*')])
        if not path:
            return
        try:
            df = pd.read_excel(path)
            col = df.columns[0]
            raw_codes = [str(c).strip() for c in df[col].dropna().tolist()]
            raw_codes = [c for c in raw_codes if c]
            total_found = len(raw_codes)

            seen = set()
            unique_codes = []
            for c in raw_codes:
                if c not in seen:
                    seen.add(c)
                    unique_codes.append(c)
            duplicates_removed = total_found - len(unique_codes)

            self.app.loaded_codes = unique_codes
            count = len(unique_codes)

            self.file_label.config(text=os.path.basename(path), fg=TEXT)
            if duplicates_removed > 0:
                msg = (f'Found {total_found} codes, '
                       f'{duplicates_removed} duplicates removed, '
                       f'{count} unique codes to process.')
                self.code_count_label.config(text=msg, fg=WARNING)
            else:
                self.code_count_label.config(
                    text=f'{count} unique codes loaded (no duplicates).',
                    fg=SUCCESS)

            self.preview_text.config(state='normal')
            self.preview_text.delete('1.0', 'end')
            for code in self.app.loaded_codes[:50]:
                self.preview_text.insert('end', code + '\n')
            if count > 50:
                self.preview_text.insert('end', f'\n... and {count - 50} more')
            self.preview_text.config(state='disabled')

        except Exception as e:
            messagebox.showerror('Error', f'Failed to read file:\n{e}')

    def process_and_next(self):
        if not self.app.loaded_codes:
            messagebox.showwarning('Warning', 'Please upload a codes file first.')
            return

        stone_count = self.app.db.get_count('stone_names')
        if stone_count == 0:
            messagebox.showwarning('Warning',
                                   'Stone Name library is empty. Please upload libraries first.')
            return

        stone_lookup = self.app.db.get_lookup('stone_names')
        polish_lookup = self.app.db.get_lookup('polishing_types')
        shape_lookup = self.app.db.get_lookup('shapes')
        origin_lookup = self.app.db.get_lookup('origins')

        results, skipped = decode_codes(
            self.app.loaded_codes,
            stone_lookup, polish_lookup, shape_lookup, origin_lookup
        )
        self.app.decoded_results = results
        self.app.skipped_codes = skipped

        if not results and not skipped:
            messagebox.showwarning('Warning',
                                   'No codes were processed. Check your input file.')
            return

        if not results:
            messagebox.showwarning('Warning',
                                   f'No codes matched the stone library. '
                                   f'{len(skipped)} codes skipped. You can still export skipped codes.')

        self.app.show('ExportScreen')


# --- Screen 3: Export ---

class ExportScreen(tk.Frame):
    def __init__(self, parent, app):
        super().__init__(parent, bg=BG)
        self.app = app

        make_header(self, app)
        make_footer(self)

        make_title(self, 'Export Results')
        make_subtitle(self, 'Review and export the decoded stone codes.')

        self.summary_frame = tk.Frame(self, bg=CARD_BG, relief='groove', bd=1)
        self.summary_frame.pack(padx=40, pady=10, fill='x', ipady=10)

        self.summary_label = tk.Label(self.summary_frame, text='',
                                       font=('Segoe UI', 11), bg=CARD_BG, fg=TEXT)
        self.summary_label.pack(pady=10)

        preview_frame = tk.Frame(self, bg=BG)
        preview_frame.pack(fill='both', expand=True, padx=40, pady=5)

        cols = ('Code', 'Attribute', 'Value')
        self.tree = ttk.Treeview(preview_frame, columns=cols, show='headings',
                                  height=12)
        for col in cols:
            self.tree.heading(col, text=col)
        self.tree.column('Code', width=160)
        self.tree.column('Attribute', width=130)
        self.tree.column('Value', width=200)

        scrollbar = ttk.Scrollbar(preview_frame, orient='vertical',
                                   command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        self.tree.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')

        btn_frame = tk.Frame(self, bg=BG)
        btn_frame.pack(pady=12)

        make_button(btn_frame, '⬅  Libraries',
                    lambda: self.app.show('LibraryScreen'),
                    color=TEXT_LIGHT, width=14).pack(side='left', padx=6)
        make_button(btn_frame, 'Export Decoded Codes',
                    self.export, color=SUCCESS, width=20).pack(side='left', padx=6)
        make_button(btn_frame, 'Export Skipped Codes',
                    self.export_skipped, color=WARNING, width=20).pack(side='left', padx=6)
        make_button(btn_frame, 'Close App',
                    self.app.on_close, color=DANGER, width=12).pack(side='left', padx=6)

    def on_show(self):
        results = self.app.decoded_results
        skipped = self.app.skipped_codes
        total_codes = len(self.app.loaded_codes)
        matched = len(results)
        skipped_count = len(skipped)

        self.summary_label.config(
            text=(f'Total codes: {total_codes}   |   '
                  f'Matched: {matched}   |   '
                  f'Skipped: {skipped_count}'))

        self.tree.delete(*self.tree.get_children())
        for r in results:
            attrs = [
                ('Stone name', r['stone_name']),
                ('Polishing', r['polishing']),
                ('Shape', r['shape']),
                ('Colour', r['colour']),
                ('Origin', r['origin']),
            ]
            for i, (attr, val) in enumerate(attrs):
                code_display = r['code'] if i == 0 else ''
                self.tree.insert('', 'end', values=(code_display, attr, val))

    def _styled_workbook(self, sheet_title, headers, rows, col_widths,
                         highlight_first_col=True):
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = sheet_title

        header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill('solid', fgColor='2563EB')
        header_align = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin', color='D0D0D0'),
            right=Side(style='thin', color='D0D0D0'),
            top=Side(style='thin', color='D0D0D0'),
            bottom=Side(style='thin', color='D0D0D0'),
        )

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        code_font = Font(name='Consolas', size=10, bold=True, color='1E293B')
        body_font = Font(name='Segoe UI', size=10, color='1E293B')
        code_fill = PatternFill('solid', fgColor='F0F4F8')

        for row_idx, row_values in enumerate(rows, 2):
            for col_idx, val in enumerate(row_values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                if col_idx == 1:
                    cell.font = code_font
                    if highlight_first_col and val:
                        cell.fill = code_fill
                else:
                    cell.font = body_font
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center')

        for col_idx, width in enumerate(col_widths, 1):
            ws.column_dimensions[chr(ord('A') + col_idx - 1)].width = width

        return wb

    def export(self):
        if not self.app.decoded_results:
            messagebox.showwarning('Warning', 'No decoded results to export.')
            return

        path = filedialog.asksaveasfilename(
            title='Save decoded codes file',
            defaultextension='.xlsx',
            filetypes=[('Excel files', '*.xlsx')],
            initialfile='Decoded_Codes.xlsx')
        if not path:
            return

        try:
            rows = []
            for r in self.app.decoded_results:
                attrs = [
                    ('Stone name', r['stone_name']),
                    ('Polishing', r['polishing']),
                    ('Shape', r['shape']),
                    ('Colour', r['colour']),
                    ('Origin', r['origin']),
                ]
                polishing = r['polishing'] if r['polishing'] else 'ROUGH'
                name = f"{r['stone_name']} {polishing}"
                for i, (attr, val) in enumerate(attrs):
                    rows.append((
                        r['code'] if i == 0 else '',
                        attr,
                        val if val else '',
                        name if i == 0 else '',
                    ))

            wb = self._styled_workbook(
                'Decoded Codes',
                ['Codes', 'Attribute', 'Value', 'Name'],
                rows,
                col_widths=[22, 16, 24, 25],
            )
            wb.save(path)
            messagebox.showinfo('Success',
                                f'Exported {len(self.app.decoded_results)} codes to:\n{path}')

        except Exception as e:
            messagebox.showerror('Error', f'Export failed:\n{e}')

    def export_skipped(self):
        if not self.app.skipped_codes:
            messagebox.showinfo('Nothing to export',
                                'There are no skipped codes to export.')
            return

        path = filedialog.asksaveasfilename(
            title='Save skipped codes file',
            defaultextension='.xlsx',
            filetypes=[('Excel files', '*.xlsx')],
            initialfile='Skipped_Codes.xlsx')
        if not path:
            return

        try:
            rows = [(s['code'], s['reason']) for s in self.app.skipped_codes]
            wb = self._styled_workbook(
                'Skipped Codes',
                ['Skipped Codes', 'Reason'],
                rows,
                col_widths=[24, 40],
                highlight_first_col=False,
            )
            wb.save(path)
            messagebox.showinfo('Success',
                                f'Exported {len(self.app.skipped_codes)} skipped codes to:\n{path}')

        except Exception as e:
            messagebox.showerror('Error', f'Export failed:\n{e}')


if __name__ == '__main__':
    app = App()
    app.mainloop()
